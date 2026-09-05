# -*- coding: utf-8 -*-
"""
seal_guide.py  船外机油封/O圈/齿轮箱密封套件 外勤带货推荐
数据源:
  oemkb.db            megazip 雅马哈 OEM 树, 全族聚合定位油封部位与适配马力
  suzuki_eparts.db    印尼铃木官方 eparts, DF 四冲; 铃木齿轮箱按 SEALING KIT 套件卖
  sales_data.json     Odoo 全员销售(复购/客户数, 含"剔除王明营156特例"口径)
  Odoo product.product 实时在手/预测/中文名
部位: 曲轴油封 / 驱动轴油封 / 车叶轴油封 / 齿轮箱密封套件(铃木) / O型圈
只读取数, 不写 Odoo。
"""
import sys, io, os, re, sqlite3, json
sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding="utf-8")
from collections import defaultdict

TOOL = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = r"C:\Users\guobi\Doubao\chats\2026-09-02\new-chat-1"
OEMKB = os.path.join(TOOL, "oemkb.db")
SUZDB = os.path.join(TOOL, "suzuki_eparts.db")
SALES = os.path.join(TOOL, "sales_data.json")
WANG_PARTNER = 156

def mega_to_compact(pn):
    if not pn: return pn
    pn = pn.strip().upper()
    m = re.match(r"^(\d{3})-(\d{2}[A-Z0-9]+?)-(\d{2})-00$", pn)
    if m:
        a, b, c = m.groups()
        return a + b[:2] + "-" + b[2:] + c
    return re.sub(r"-00$", "", pn)

def norm(code):
    """件号归一单一真相源 cn_stock.norm（保留材质/W/H）。本脚本为旧版，新油封推荐任务请用 build_seal_rec。"""
    if not code: return ""
    import cn_stock
    return cn_stock.norm(code)

def is_pure_seal(desc):
    d = (desc or "").lower()
    return d.startswith("oil seal") and all(x not in d for x in
           ["housing", "cover", "gasket", "protector"])

def newentry():
    return {"loc": set(), "hp": set(), "brand": set(), "mega": set(), "spec": set()}

# ---------- 雅马哈(megazip) ----------
def build_yamaha(PM):
    con = sqlite3.connect(OEMKB); cur = con.cursor()
    fams = cur.execute("SELECT url,brand,hp FROM family").fetchall()
    def inum(x):
        m = re.match(r"\d+", str(x or "")); return int(m.group()) if m else 0
    for furl, brand, hp in fams:
        mm = re.match(r"[fFlL]*0*(\d+)", str(hp or ""))
        hpnum = int(mm.group(1)) if mm else None
        if not (str(hp or "").lower().startswith("f") and hpnum and 60 <= hpnum <= 300):
            continue
        rec = cur.execute("""SELECT s.name,p.item,p.part_no,p.desc FROM part p
                  JOIN section s ON p.sec_url=s.url JOIN unit u ON s.unit_url=u.url
                  WHERE u.family_url=?""", (furl,)).fetchall()
        secmap = defaultdict(list)
        for sn, item, pn, ds in rec:
            secmap[sn].append((inum(item), pn, ds or ""))
        for sn, items in secmap.items():
            crank = sn.lower().startswith("crank")
            d_items = [it for it, pn, ds in items if "drive shaft" in ds.lower()]
            p_items = [it for it, pn, ds in items if "shaft, propeller" in ds.lower()]
            for it, pn, ds in items:
                dl = ds.lower()
                is_oring = ("o-ring" in dl or "o ring" in dl
                            or (pn or "").upper().replace("-", "").startswith("932"))
                if not (is_pure_seal(ds) or is_oring): continue
                cp = mega_to_compact(pn)
                if is_oring: loc = "O型圈"
                elif crank: loc = "曲轴油封"
                elif d_items or p_items:
                    if d_items and not p_items: loc = "驱动轴油封"
                    elif p_items and not d_items: loc = "车叶轴油封"
                    else:
                        nd = min((abs(it - x) for x in d_items), default=9999)
                        npp = min((abs(it - x) for x in p_items), default=9999)
                        loc = "驱动轴油封" if nd <= npp else "车叶轴油封"
                else: loc = "下箱体油封"
                e = PM[cp]; e["loc"].add(loc); e["hp"].add(hpnum)
                e["brand"].add("yamaha"); e["mega"].add((pn or "").upper()); e["spec"].add(ds.strip())
    con.close()

# ---------- 铃木(印尼 eparts) ----------
def build_suzuki(PM):
    if not os.path.exists(SUZDB): return
    con = sqlite3.connect(SUZDB); cur = con.cursor()
    rows = cur.execute("""SELECT v.name,f.name,p.part_no,p.name FROM part p
        JOIN figure f ON p.figure_id=f.id JOIN vehicle v ON f.vehicle_id=v.id
        WHERE v.name LIKE 'DF%'""").fetchall()
    for vn, fn, pn, pname in rows:
        fnu = (fn or "").upper(); pnl = (pname or "").lower()
        hps = [int(x) for x in re.findall(r"\d{2,3}", vn or "") if 60 <= int(x) <= 300]
        code = re.sub(r"-\d{3}$", "", (pn or "").strip())
        loc = None
        if "CRANKSHAFT" in fnu and pnl.startswith("oil seal"):
            loc = "曲轴油封"
        elif "GEAR CASE" in fnu and ("sealing kit" in pnl or "kit, gear case" in pnl):
            loc = "齿轮箱密封套件"
        elif "GEAR CASE" in fnu and pnl.startswith("oil seal"):
            loc = "齿轮箱油封"
        elif "DRIVE SHAFT" in fnu and pnl.startswith("oil seal"):
            loc = "驱动轴油封"
        if not loc: continue
        e = PM[code]; e["loc"].add(loc); e["brand"].add("suzuki"); e["spec"].add(pname.strip())
        for h in hps: e["hp"].add(h)
    con.close()

def build_sales():
    sd = json.load(open(SALES, encoding="utf-8"))
    om = {o["id"]: o for o in sd["orders"]}
    S = defaultdict(lambda: {"q": 0.0, "od": set(), "pc": set(),
                             "q2": 0.0, "od2": set(), "pc2": set(), "cn": ""})
    for l in sd["lines"]:
        m = re.search(r"\[([^\]]+)\]", l.get("name", ""))
        if not m: continue
        code = norm(m.group(1)); o = om.get(l.get("order"))
        pid = o.get("partner") if o else None
        a = S[code]; qty = l.get("qty", 0) or 0
        a["q"] += qty; a["od"].add(l.get("order")); a["pc"].add(pid)
        a["cn"] = re.sub(r"^\[[^\]]+\]\s*", "", l.get("name", ""))
        if str(pid) != str(WANG_PARTNER):
            a["q2"] += qty; a["od2"].add(l.get("order")); a["pc2"].add(pid)
    return S

def build_stock():
    from odoo.client import OdooClient
    from config import ODOO_URL, ODOO_DB, ODOO_UID, ODOO_PWD
    c = OdooClient(ODOO_URL, ODOO_DB, ODOO_UID, ODOO_PWD)
    prods = c.search_read("product.product", [],
        fields=["default_code", "qty_available", "virtual_available", "name"], limit=50000)
    ST = {}
    for p in prods:
        k = norm(p.get("default_code"))
        if k: ST[k] = {"onhand": p.get("qty_available", 0),
                       "forecast": p.get("virtual_available", 0), "name": p.get("name", "")}
    return ST

LOC_ORDER = ["曲轴油封", "驱动轴油封", "车叶轴油封", "下箱体油封", "齿轮箱密封套件", "齿轮箱油封", "O型圈"]
def loc_rank(locset):
    for o in LOC_ORDER:
        if o in locset: return o
    return "/".join(sorted(locset))

def main():
    print("[1/5] 雅马哈 OEM 聚合 ..."); PM = defaultdict(newentry); build_yamaha(PM)
    print("[2/5] 铃木 DF 聚合 ..."); build_suzuki(PM)
    print("   油封/套件/O圈件号数:", len(PM))
    print("[3/5] 销售聚合 ..."); S = build_sales()
    print("[4/5] Odoo 库存 ..."); ST = build_stock(); print("   产品档案:", len(ST))

    master = []
    for code, e in PM.items():
        st = ST.get(norm(code), {"onhand": None, "forecast": None, "name": ""})
        sl = S.get(norm(code)); hps = sorted(e["hp"])
        spec = sorted(e["spec"])[0] if e["spec"] else ""
        master.append({"code": code, "loc": loc_rank(e["loc"]), "hps": hps, "nhp": len(hps),
            "brand": "/".join(sorted(e["brand"])),
            "name": st["name"] or spec, "spec": spec,
            "onhand": st["onhand"], "forecast": st["forecast"],
            "q": sl["q"] if sl else 0, "npc": len(sl["pc"]) if sl else 0,
            "npc2": len(sl["pc2"]) if sl else 0})
    master = [m for m in master if m["nhp"] > 0]
    def rec_level(m):
        if m["loc"] == "齿轮箱密封套件":
            return "铃木套件(整套)"
        if m["onhand"] is not None and m["onhand"] <= 0: return "需补货"
        if m["npc2"] >= 2: return "走量必带"
        if m["nhp"] >= 4 and (m["onhand"] or 0) > 0: return "通用常备"
        if m["loc"] == "曲轴油封": return "补充品类(曲轴)"
        return "按需"
    for m in master: m["rec"] = rec_level(m)
    recorder = {"走量必带": 0, "通用常备": 1, "铃木套件(整套)": 2, "需补货": 3,
                "补充品类(曲轴)": 4, "按需": 5}
    master.sort(key=lambda m: (LOC_ORDER.index(m["loc"]) if m["loc"] in LOC_ORDER else 9,
                               recorder.get(m["rec"], 9), -m["nhp"], -m["q"]))
    longrows = []
    for m in master:
        for hp in m["hps"]:
            longrows.append([m["brand"], hp, m["loc"], m["code"], m["name"], m["onhand"],
                             m["forecast"], m["q"], m["npc"], m["npc2"], m["rec"]])
    longrows.sort(key=lambda r: (r[1], LOC_ORDER.index(r[2]) if r[2] in LOC_ORDER else 9, r[3]))

    sold = []
    for code, a in S.items():
        if re.match(r"^(9310|93101|93102|93106|93210|26-|6H1|0928|25700)", code):
            e = PM.get(code); st = ST.get(norm(code), {"onhand": None, "forecast": None})
            sold.append([code, a["cn"], loc_rank(e["loc"]) if e else "(库外)",
                         "/".join(map(str, sorted(e["hp"]))) if e else "",
                         a["q"], len(a["pc"]), len(a["pc2"]), st["onhand"], st["forecast"]])
    sold.sort(key=lambda r: -r[4])

    print("[5/5] 写 Excel ...")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = openpyxl.Workbook()
    hf = Font(bold=True, color="FFFFFF", size=10); hfill = PatternFill("solid", fgColor="2F5597")
    thin = Border(*[Side(style="thin", color="D9D9D9")] * 4)
    def sheet(ws, headers, data, widths, wrapcols=()):
        ws.append(headers)
        for c in ws[1]:
            c.font = hf; c.fill = hfill; c.alignment = Alignment(horizontal="center", vertical="center"); c.border = thin
        for row in data: ws.append(row)
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        for r in ws.iter_rows(min_row=2):
            for c in r:
                c.border = thin; c.font = Font(size=9)
                if c.column in wrapcols: c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
    ws0 = wb.active; ws0.title = "口径说明"
    notes = [
        ["船外机油封/O圈/齿轮箱密封套件 外勤带货推荐 — 口径说明"], [],
        ["1. 部位定义", "曲轴油封=发动机Crankshaft组; 驱动轴油封=齿轮箱Drive Shaft处; 车叶轴油封=螺旋桨轴Propeller处(雅马哈同组按零件相邻关系定性)"],
        ["2. 铃木结构差异", "铃木DF四冲齿轮箱不单独卖驱动轴/车叶轴油封, 只卖整套【齿轮箱密封套件 25700-xx】(内含全部油封+O圈); 曲轴油封按单件 09283-xxxx 卖, 名称括号内是油封尺寸(内径X外径X厚)"],
        ["3. 适配马力", "雅马哈来自 megazip 全族聚合(F60-F300四冲, 跨年份取并集); 铃木来自印尼官方eparts(suzuki_eparts.db)"],
        ["4. 库存口径", "Odoo product.product 实时在手/预测; 在手None=Odoo无此档案(需建档/采购), 0=有档案但无货"],
        ["5. 销量口径", "全员销售(6业务员,2026.6-9); 复购=同客户不同订单; 同一订单多数量只算一次"],
        ["6. 王明营特例", "宁德王明营partner156那张7100为零库存起步备货、之后仅补67元, 不代表普遍需求, 单列[剔王客户数]排序"],
        ["7. 推荐等级", "走量必带=剔王≥2客户; 通用常备=覆盖≥4马力且有货; 补充品类(曲轴)=销售空白但维修必换; 需补货=在手0; 铃木套件=齿轮箱密封整套"],
    ]
    for r in notes: ws0.append(r)
    ws0.column_dimensions["A"].width = 16; ws0.column_dimensions["B"].width = 115
    ws0["A1"].font = Font(bold=True, size=12)
    for r in ws0.iter_rows(min_row=3):
        r[0].font = Font(bold=True, size=9); r[1].alignment = Alignment(wrap_text=True)
    ws1 = wb.create_sheet("油封件号主数据")
    sheet(ws1, ["件号", "部位", "品牌", "适配马力", "覆盖马力数", "Odoo中文名/规格",
                "在手", "预测", "全员销量", "客户数", "剔王客户", "推荐"],
          [[m["code"], m["loc"], m["brand"], "/".join(map(str, m["hps"])), m["nhp"], m["name"],
            m["onhand"], m["forecast"], m["q"], m["npc"], m["npc2"], m["rec"]] for m in master],
          [16, 14, 9, 24, 8, 36, 7, 7, 8, 7, 8, 15], wrapcols=(4, 6))
    ws2 = wb.create_sheet("机型x部位对照")
    sheet(ws2, ["品牌", "马力", "部位", "件号", "中文名/规格", "在手", "预测",
                "销量", "客户", "剔王", "推荐"], longrows,
          [9, 6, 14, 16, 36, 7, 7, 7, 6, 6, 15], wrapcols=(5,))
    ws3 = wb.create_sheet("已售密封件验证")
    sheet(ws3, ["件号", "中文名", "部位", "适配马力", "销量", "客户", "剔王", "在手", "预测"], sold,
          [16, 30, 14, 22, 7, 6, 6, 7, 7], wrapcols=(2, 4))
    out = os.path.join(OUT_DIR, "2026.9.5油封O圈带货推荐-机型部位对照.xlsx")
    wb.save(out)
    rb = openpyxl.load_workbook(out)
    print("SAVED", out)
    for ws in rb.worksheets: print("  sheet", ws.title, "行", ws.max_row)
    for lvl in ["走量必带", "需补货", "铃木套件(整套)"]:
        print("===", lvl, "===")
        for m in master:
            if m["rec"] == lvl:
                print(" ", m["code"], m["loc"], m["brand"], "在手", m["onhand"], "HP", m["hps"])

if __name__ == "__main__":
    main()
