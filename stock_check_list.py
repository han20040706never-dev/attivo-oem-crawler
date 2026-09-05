# -*- coding: utf-8 -*-
"""
任意 Excel 清单一键核“中国仓实时库存” —— 高频复用工具。

用法示例:
    python stock_check_list.py 清单.xlsx
    python stock_check_list.py 清单.xlsx --col B --sheets Sheet1,Sheet2
    python stock_check_list.py 清单.xlsx --fresh --out 结果.xlsx

说明:
- 用 openpyxl load_workbook(data_only=True, read_only=True) 读；
  默认遍历所有 sheet，--sheets 指定（逗号分隔）。
- 件号列默认第1列(A)，--col 可改列字母。
- 首行若任一单元格含 “OEM/件号/编号/ID/品类” 视为表头跳过。
- 收集非空件号（去重、保序）。若存在第2列当品类、能找到“数量/需求”列则一并记录。
- sys.path.insert(0, TOOL) 后 import cn_stock；
  默认 cn_stock.qty_with_verify(codes, verify_zero=True)（缓存优先、缓存判0自动实时复核）；
  加 --fresh 时直接 cn_stock.live_qty(codes) 全实时。
- 控制台紧凑输出：按 sheet 分组打印 “件号 | 品类 | 中国仓实时 | 来源cache/live”；
  末尾汇总 唯一件数、有货数、缺货(<=0)件号清单。
- --out 时用 openpyxl 新建结果工作簿（绝不覆盖原 xlsx），
  表头 [来源sheet, 件号, 品类, 中国仓实时, 数据来源, 状态]，状态=有货/缺货；
  加首行冻结 + 自动列宽即可，不要复杂样式。
- 件号归一用 cn_stock.norm；同一 norm 合并数量；任何单个 sheet 异常跳过并 WARN 不中断整体。
"""

import sys
import os
import io
import argparse

# stdout 用 utf-8 包装
sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding='utf-8')

TOOL = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, TOOL)

# 延迟导入
try:
    import cn_stock
except ImportError:
    cn_stock = None


HEADER_KEYWORDS = ('OEM', '件号', '编号', 'ID', '品类')


def _is_header_row(row_vals):
    """判断首行是否为表头。"""
    for val in row_vals:
        if val is None:
            continue
        s = str(val).strip()
        if any(kw in s for kw in HEADER_KEYWORDS):
            return True
    return False


def _col_letter_to_idx(col_letter):
    """列字母转 0-based 索引，如 A->0, B->1, AA->26。"""
    idx = 0
    for ch in col_letter.upper():
        idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx - 1


def _read_sheet(sheet, code_col_idx, need_qty_col):
    """读取单个 sheet，返回 (codes_list, qty_map, category_map)。
    codes_list: 保序去重后的件号列表
    qty_map: {norm_code: 需求数量}（若找到数量列）
    category_map: {norm_code: 品类}
    """
    codes = []
    seen = set()
    qty_map = {}
    category_map = {}

    # 先找数量列（若需要）
    qty_col_idx = None
    if need_qty_col:
        # 读取第一行找表头
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if first_row:
            for i, val in enumerate(first_row):
                if val and '数量' in str(val) or (val and '需求' in str(val)):
                    qty_col_idx = i
                    break

    start_row = 2 if _is_header_row(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])) else 1

    for row in sheet.iter_rows(min_row=start_row, values_only=True):
        if code_col_idx >= len(row):
            continue
        code_raw = row[code_col_idx]
        if code_raw is None:
            continue
        code_str = str(code_raw).strip()
        if not code_str:
            continue
        try:
            norm = cn_stock.norm(code_str)
        except Exception:
            norm = code_str
        if norm in seen:
            continue
        seen.add(norm)
        codes.append(norm)

        # 品类（第2列，即索引1）
        if len(row) > 1 and row[1] is not None:
            category_map[norm] = str(row[1]).strip()

        # 数量
        if qty_col_idx is not None and qty_col_idx < len(row) and row[qty_col_idx] is not None:
            try:
                qty_map[norm] = float(row[qty_col_idx])
            except (ValueError, TypeError):
                pass

    return codes, qty_map, category_map


def main():
    parser = argparse.ArgumentParser(description='Excel 清单核中国仓实时库存')
    parser.add_argument('xlsx', help='Excel 文件路径')
    parser.add_argument('--col', default='A', help='件号列字母，默认 A')
    parser.add_argument('--sheets', default=None, help='指定 sheet 名，逗号分隔；默认全部')
    parser.add_argument('--fresh', action='store_true', help='强制实时查询，不使用缓存')
    parser.add_argument('--out', default=None, help='输出结果 xlsx 路径')
    args = parser.parse_args()

    if cn_stock is None:
        print("FAIL: 无法导入 cn_stock 模块，请确认同目录存在 cn_stock.py")
        sys.exit(1)

    xlsx_path = args.xlsx
    if not os.path.isfile(xlsx_path):
        print(f"FAIL: Excel 文件不存在 {xlsx_path}")
        sys.exit(1)

    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    except Exception as e:
        print(f"FAIL: 打开 Excel 失败 {xlsx_path} | {str(e)[:80]}")
        sys.exit(1)

    # 确定要处理的 sheet
    if args.sheets:
        sheet_names = [s.strip() for s in args.sheets.split(',') if s.strip()]
    else:
        sheet_names = wb.sheetnames

    code_col_idx = _col_letter_to_idx(args.col)

    all_codes = []          # 全局保序去重
    all_seen = set()
    sheet_data = []         # [(sheet_name, codes, qty_map, category_map)]

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            print(f"WARN: sheet 不存在，跳过 {sheet_name}")
            continue
        ws = wb[sheet_name]
        try:
            codes, qty_map, category_map = _read_sheet(ws, code_col_idx, need_qty_col=True)
            sheet_data.append((sheet_name, codes, qty_map, category_map))
            for c in codes:
                if c not in all_seen:
                    all_seen.add(c)
                    all_codes.append(c)
        except Exception as e:
            print(f"WARN: 读取 sheet {sheet_name} 失败，跳过 | {str(e)[:80]}")
            continue

    if not all_codes:
        print("FAIL: 未收集到任何有效件号")
        sys.exit(1)

    # 查询库存
    try:
        if args.fresh:
            stock_map = cn_stock.live_qty(all_codes)
            source_tag = 'live'
            src_map = {c: 'live' for c in stock_map}
        else:
            raw = cn_stock.qty_with_verify(all_codes, verify_zero=True)
            # qty_with_verify 返回 {code:{'qty':,'src':}}，统一规整成纯数值 stock_map + 逐件 src_map
            src_map = {k: (v.get('src', 'cache') if isinstance(v, dict) else 'cache') for k, v in raw.items()}
            stock_map = {k: (v.get('qty', 0) if isinstance(v, dict) else (v or 0)) for k, v in raw.items()}
            source_tag = 'cache'
    except Exception as e:
        print(f"FAIL: 库存查询失败 | {str(e)[:80]}")
        sys.exit(1)

    # 控制台输出
    print("\n=== 库存核对结果 ===")
    for sheet_name, codes, qty_map, category_map in sheet_data:
        print(f"\n--- {sheet_name} ---")
        for code in codes:
            qty = stock_map.get(code, 0)
            cat = category_map.get(code, '')
            src = src_map.get(code, source_tag)
            print(f"{code} | {cat} | {qty} | {src}")

    # 汇总
    unique_count = len(all_codes)
    in_stock = [c for c in all_codes if stock_map.get(c, 0) > 0]
    out_stock = [c for c in all_codes if stock_map.get(c, 0) <= 0]
    print(f"\n=== 汇总 ===")
    print(f"唯一件数: {unique_count}")
    print(f"有货数: {len(in_stock)}")
    print(f"缺货数: {len(out_stock)}")
    if out_stock:
        print(f"缺货件号: {', '.join(out_stock)}")

    # 输出 Excel
    if args.out:
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            wb_out = Workbook()
            ws_out = wb_out.active
            ws_out.title = '库存核对结果'
            headers = ['来源sheet', '件号', '品类', '中国仓实时', '数据来源', '状态']
            ws_out.append(headers)

            for sheet_name, codes, qty_map, category_map in sheet_data:
                for code in codes:
                    qty = stock_map.get(code, 0)
                    cat = category_map.get(code, '')
                    status = '有货' if qty > 0 else '缺货'
                    ws_out.append([sheet_name, code, cat, qty, src_map.get(code, source_tag), status])

            # 冻结首行
            ws_out.freeze_panes = 'A2'
            # 自动列宽（粗略）
            for col_idx in range(1, len(headers) + 1):
                max_len = len(headers[col_idx - 1])
                for row in ws_out.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
                    for val in row:
                        if val is not None:
                            max_len = max(max_len, len(str(val)))
                ws_out.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

            wb_out.save(args.out)
            print(f"\n结果已保存: {args.out}")
        except Exception as e:
            print(f"FAIL: 输出 Excel 失败 | {str(e)[:80]}")
            sys.exit(1)

    print("\nOK 处理完成")


if __name__ == '__main__':
    main()
