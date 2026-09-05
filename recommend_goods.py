# -*- coding: utf-8 -*-
"""外勤带货综合推荐：需求清单 ∪ 全员销售小件 ∪ 油封OEM主数据，结合Odoo在手，只四冲。"""
import sys, io, os, re, json
from collections import defaultdict
sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding="utf-8")
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TOOL = r"C:\Users\guobi\Doubao\chats\2026-08-16\new-chat\attivo-odoo-tools"
OUTDIR = r"C:\Users\guobi\Doubao\chats\2026-09-02\new-chat-1"
SEAL_X = os.path.join(OUTDIR, "2026.9.5油封O圈带货推荐-机型部位对照.xlsx")
WANG = 156

def norm(c):
    c = re.sub(r"\(.*?\)", "", (c or "").upper())
    # W/H=With Housing 是不同物料，绝不能合并；只归一格式尾缀
    c = re.sub(r"-(OIL|ZN|AL|O|MARTYR|RIKEN|COPPER|BEST-[A-Z0-9]+)$", "", c)
    c = re.sub(r"-0{2,3}$", "", c)
    return c.strip()

# ---------- 1. 需求清单 ----------
dem = json.load(open(os.path.join(TOOL, "_demand.json"), encoding="utf-8"))
D = {}
for sheet, items in dem.items():
    for d in items:
        code = norm(str(d.get("OEM ID", "")))
        if not code: continue
        D[code] = {"cat": str(d.get("品类", "")), "brand": str(d.get("品牌", "")),
                   "hp4": str(d.get("4冲程", "")), "plan": d.get("需求数量"), "sheet": sheet}

# ---------- 2. 销售聚合(全员, 剔王明营, 复购客户) ----------
sd = json.load(open(os.path.join(TOOL, "sales_data.json"), encoding="utf-8"))
om = {o["id"]: o for o in sd["orders"]}
S = defaultdict(lambda: {"q": 0.0, "od": set(), "pc": set(), "q2": 0.0, "pc2": set(),
                         "od2": set(), "cn": "", "po": defaultdict(set)})
for l in sd["lines"]:
    m = re.search(r"\[([^\]]+)\]", l.get("name", ""))
    if not m: continue
    code = norm(m.group(1)); o = om.get(l.get("order")); pid = o.get("partner") if o else None
    a = S[code]; q = l.get("qty", 0) or 0
    a["q"] += q; a["od"].add(l.get("order")); a["pc"].add(pid)
    a["cn"] = re.sub(r"^\[[^\]]+\]\s*", "", l.get("name", ""))
    if str(pid) != str(WANG):
        a["q2"] += q; a["pc2"].add(pid); a["od2"].add(l.get("order"))
    if o: a["po"][pid].add(l.get("order"))
def repn(a): return sum(1 for p, os_ in a["po"].items() if len(os_) >= 2)

# ---------- 3. 油封OEM主数据(部位/适配马力) ----------
SEAL = {}
if os.path.exists(SEAL_X):
    wb = openpyxl.load_workbook(SEAL_X, data_only=True)
    ws = wb["油封件号主数据"]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[0]: continue
        SEAL[norm(str(r[0]))] = {"loc": r[1], "brand": r[2], "hp": str(r[3] or ""), "nm": r[5] or ""}

# ---------- 4. Odoo 在手(只算中国仓WH2/库位根1563) + 磁盘缓存(6h, --fresh强刷;省重复全量拉取) ----------
import time as _time
_CACHE_STOCK = os.path.join(TOOL, "_cn_stock_cache.json")
ST = None; prods = []; cnids = set()
if "--fresh" not in sys.argv and os.path.exists(_CACHE_STOCK):
    try:
        _sc = json.load(open(_CACHE_STOCK, encoding="utf-8"))
        if _time.time() - _sc.get("ts", 0) < 6 * 3600:
            ST = _sc["ST"]; prods = list(range(_sc.get("prods", 0)))
            print(f"用中国仓库存缓存({_sc.get('prods')}产品/{_time.strftime('%m-%d %H:%M', _time.localtime(_sc['ts']))})，--fresh强刷")
    except Exception:
        ST = None
if ST is None:
    sys.path.insert(0, TOOL)
    from config import ODOO_URL, ODOO_DB, ODOO_UID, ODOO_PWD
    from odoo.client import OdooClient
    cli = OdooClient(ODOO_URL, ODOO_DB, ODOO_UID, ODOO_PWD)
    CN_ROOT = 1563
    _cnl = cli.search_read("stock.location", [("id","child_of",CN_ROOT),("usage","=","internal")], ["id"], limit=5000)
    cnids = {l["id"] for l in _cnl}
    _qu = cli.search_read("stock.quant", [], ["product_id","location_id","quantity"], limit=200000)
    CNQ = defaultdict(float)
    for q in _qu:
        loc = q["location_id"][0] if isinstance(q.get("location_id"), (list,tuple)) else q.get("location_id")
        if loc in cnids:
            pid = q["product_id"][0] if isinstance(q.get("product_id"), (list,tuple)) else q.get("product_id")
            CNQ[pid] += q.get("quantity") or 0
    prods = cli.search_read("product.product", [], ["default_code", "qty_available", "name"], limit=60000)
    ST = {}
    for p in prods:
        dc = p.get("default_code")
        if not dc: continue
        k = norm(dc); cn = round(CNQ.get(p["id"], 0.0), 1); gl = p.get("qty_available") or 0
        if k in ST: ST[k]["cn"] += cn; ST[k]["gl"] += gl
        else: ST[k] = {"cn": cn, "gl": gl, "name": p.get("name", "")}
    print("Odoo 产品档案", len(prods), "中国仓库位", len(cnids))
    try:
        json.dump({"ts": _time.time(), "prods": len(prods), "ST": ST},
                  open(_CACHE_STOCK, "w", encoding="utf-8"), ensure_ascii=False)
    except Exception:
        pass

# ---------- 5. 品类/品牌归类 ----------
def brand_of(code, dcat, seal):
    if seal and code in SEAL: return {"yamaha": "雅马哈", "suzuki": "铃木"}.get(SEAL[code]["brand"], "")
    if code in D: return D[code]["brand"]
    if code.startswith(("931", "932", "933", "6", "703", "90", "5G", "4X", "13330K")): return "雅马哈"
    if code.startswith(("0928", "25700", "154", "165", "174", "151", "152", "573", "575", "09282")): return "铃木"
    if code.startswith(("26-", "47-", "35-", "8M", "88", "80", "8M0", "19210")): return "水星/本田"
    return ""

BIG = ("齿轮", "驱动轴", "螺旋桨轴", "燃油泵", "活塞环", "连杆")
def cov_count(hp):
    """适配马力覆盖广度：覆盖几个不同马力型号；跨马力通用件=99最广"""
    if not hp: return 0
    if "通用件" in hp or "跨马力" in hp: return 99
    return len(set(re.findall(r"\d{2,3}", str(hp))))


def bigcat(cat, cn):
    # 油封/密封套件名字里虽含"驱动轴/齿轮"，但属易损密封件，不算大件
    if any(k in cat for k in ("油封", "密封", "型圈")): return False
    t = cat + cn
    return any(k in t for k in BIG)
def is_filter(cat, cn):
    return ("滤" in cat) or ("filter" in cn.lower())
def infer_cat(code, cn):
    if "手压泵" in cn or "24360" in code: return "手压输油泵"
    if "软管" in cn or code.startswith("90890"): return "燃油软管"
    if "44352" in code or "叶轮" in cn or code.startswith("17461"): return "叶轮"
    if "W0078" in code or code.startswith("17400") or "水泵维修" in cn: return "水泵维修套件"
    if any(k in code for k in ("44341","44311")) or "水泵壳体" in cn or cn == "芯子": return "水泵壳体/芯子"
    if "24305" in code or "24304" in code or "管接头" in cn: return "油管接头"
    if "半圆键" in cn or code.startswith("90280"): return "半圆键"
    if "45371" in code or "45251" in code or "阳极" in cn or "anode" in cn.lower(): return "阳极(防腐蚀)"
    if code.startswith("933") or "bearing" in cn.lower() or "轴承" in cn: return "轴承"
    if "滤" in cn: return "滤芯类"
    if "油封" in cn or code.startswith(("931","09282","26-")): return "油封"
    if "型圈" in cn: return "O圈"
    if "活塞环" in cn: return "活塞环(大件)"
    if "连杆" in cn: return "连杆(大件)"
    if any(k in cn for k in ("螺母","锁片","圆柱销","螺塞")): return "紧固件/螺塞"
    if "齿轮" in cn: return "齿轮(大件)"
    return "其他通用件"

# ---------- 6. 合并全集 ----------
codes = set(D) | set(SEAL) | {c for c, a in S.items() if len(a["pc2"]) >= 1 or len(a["pc"]) >= 1}
rows = []
for code in codes:
    a = S.get(code); seal = SEAL.get(code); dd = D.get(code); st = ST.get(code)
    cn = (a["cn"] if a else "") or (seal["nm"] if seal else "") or (st["name"] if st else "")
    # 油封/铃木套件部位以OEM爆炸图(seal)为准，覆盖需求清单里标反的标签
    cat = (seal["loc"] if seal else "") or ((dd["cat"] if dd else "") or infer_cat(code, cn))
    brand = brand_of(code, cat, seal)
    # 4冲适配马力
    if dd and dd["hp4"] not in ("None", "无", "", None):
        hp = dd["hp4"]
    elif seal:
        hp = seal["hp"]
    else:
        hp = "通用件·跨马力" if not bigcat(cat, cn) else ""
    onhand = st["cn"] if st else None
    globaloh = st["gl"] if st else None
    q2 = a["q2"] if a else 0; pc2 = len(a["pc2"]) if a else 0
    od2 = len(a["od2"]) if a else 0; rep = repn(a) if a else 0
    qall = a["q"] if a else 0; pcall = len(a["pc"]) if a else 0
    plan = dd["plan"] if dd else None
    isbig = bigcat(cat, cn); isf = is_filter(cat, cn)
    # 覆盖广度：明确列≥5个马力=广；"跨马力通用"是兜底标签，须是易损品类且真卖过才算(防杂件注水)
    cov = cov_count(hp)
    _易损 = any(k in cat for k in ("油封", "型圈", "密封", "叶轮", "水泵", "断电器", "断路器", "碳刷", "阳极", "油管", "输油泵", "接头"))
    wide = (cov >= 5 and cov != 99) or (cov == 99 and _易损 and (pc2 >= 1 or pcall >= 1))
    score = pc2 * 3 + rep * 4 + min(od2, 6) + (2 if wide else 0)
    # ===== 销售视角分级：只看"好不好卖 + 覆盖广不广"，库存只决定这次带不带得出去 =====
    if isbig:
        pr, adv = "大件·按单", "一锤子买卖不复购，客户订了再发，不用随身带"
    elif isf:
        if pc2 >= 4: pr, adv = "常备(少量)", "有几个客户拿过，但国产/假原厂竞争强，带一两个就行"
        else: pr, adv = "按单", "保养件被国产5元件占了，不主动带、要了再发"
    else:
        hot = pc2 >= 5 or rep >= 2
        reps = f"、{rep}人复购" if rep else ""
        if hot:
            pr = "⭐好卖必带"
            if onhand is None: adv = f"{pc2}个客户买过{reps}，最好卖；系统无档案，先确认能否拿货"
            elif float(onhand) == 0: adv = f"{pc2}个客户买过{reps}，好卖但中国仓现在没货、这次带不出(要先跟仓库说)"
            elif float(onhand) < 10: adv = f"{pc2}个客户买过{reps}，好卖但仓里只剩{int(float(onhand))}个，优先拿"
            else: adv = f"{pc2}个客户买过{reps}，最好卖，优先带上车"
        elif wide and (onhand or 0) and float(onhand) > 0:
            pr = "⭐广覆盖通带"
            ws = "跨马力通用" if cov == 99 else f"一件覆盖{cov}个马力机型"
            adv = f"{ws}，上门大概率碰得上，随身带1-2个保底"
        elif pc2 >= 2 or rep >= 1:
            pr, adv = "常备", f"{pc2}个客户买过{'、有人复购' if rep else ''}，带几个保底"
        elif code in SEAL:
            pr, adv = "配套带样", "油封/O圈按对应机型配套带"
        elif pcall >= 1 and pc2 == 0:
            pr, adv = "别带", "只有王明营一次性拿过、不具普遍性，不用带"
        else:
            pr, adv = "按清单", "清单内件，按计划客户带"
    # 库存提示(中国仓)
    if onhand is None: stock_note = "无档案"
    elif float(onhand) == 0: stock_note = "0·中国仓缺"
    else: stock_note = int(float(onhand))
    gnote = "" if globaloh is None else (int(globaloh) if float(globaloh) == int(globaloh) else round(globaloh, 1))
    rows.append([pr, cat or ("密封件" if code in SEAL else "其他"), code, brand,
                 hp[:80], stock_note, gnote, int(qall), pcall, int(q2), pc2, rep,
                 "" if plan is None else plan, adv, cn[:40], score])

# 排序: 优先级权重 -> score
prow = {"⭐好卖必带": 0, "⭐广覆盖通带": 1, "常备": 2, "常备(少量)": 3, "配套带样": 4,
        "按清单": 5, "大件·按单": 6, "按单": 7, "别带": 8}
rows.sort(key=lambda r: (prow.get(r[0], 9), -r[15]))

# ---------- 7. 写 Excel ----------
wb = openpyxl.Workbook()
H = Font(bold=True, color="FFFFFF", size=10); HF = PatternFill("solid", fgColor="2F5597")
bd = Border(*[Side(style="thin", color="D9D9D9")] * 4)
wrap = Alignment(wrap_text=True, vertical="center")
def style_sheet(ws, headers, widths):
    for j, h in enumerate(headers, 1):
        c = ws.cell(1, j, h); c.font = H; c.fill = HF; c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = bd
    for j, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"

ws = wb.active; ws.title = "带货推荐总表"
hdr = ["优先级", "品类/部位", "件号", "品牌", "4冲适配马力/通用性", "中国仓在手", "全球在手(参考)",
       "全员销量", "总客户", "剔王销量", "剔王客户", "复购客户", "清单计划量", "建议", "产品名", "_s"]
style_sheet(ws, hdr[:-1], [13, 13, 20, 9, 40, 10, 10, 8, 7, 8, 8, 8, 9, 30, 26])
pfill = {"⭐好卖必带": "C6EFCE", "⭐广覆盖通带": "FFEB9C", "常备": "DDEBF7"}
for r in rows:
    ws.append(r[:-1]); i = ws.max_row
    for j in range(1, 15):
        ws.cell(i, j).border = bd; ws.cell(i, j).alignment = wrap
    if r[0] in pfill:
        for j in range(1, 15): ws.cell(i, j).fill = PatternFill("solid", fgColor=pfill[r[0]])
ws.auto_filter.ref = f"A1:O{ws.max_row}"

# 结论 sheet
ws2 = wb.create_sheet("重点结论")
top = [r for r in rows if r[0].startswith("⭐")]
concl = ["【外勤带货 · 销售视角：什么好卖带什么 + 什么覆盖广带什么（全员销售剔王明营大单，只四冲）】", ""]
concl += ["一、最好卖（多客户买过/有复购），优先装包："]
for r in [x for x in rows if x[0] == "⭐好卖必带"][:12]:
    concl.append(f"  • {r[2]}（{r[3]}{r[1]}）{r[10]}客户买过/复购{r[11]}，中国仓{r[5]} — {r[13]}")
concl += ["", "二、覆盖广（一件通吃多个马力），上门大概率用得上，各带1-2："]
for r in [x for x in rows if x[0] == "⭐广覆盖通带"][:12]:
    concl.append(f"  • {r[2]}（{r[3]}{r[1]}）{str(r[4])[:48]}，中国仓{r[5]}")
concl += ["", "三、好卖但中国仓这次没货/货太少（带不出，先跟仓库打招呼）："]
_short = [x for x in rows if x[0] == "⭐好卖必带" and (x[5] in ("无档案", "0·中国仓缺") or (isinstance(x[5], (int, float)) and x[5] < 10))]
for r in _short[:10]:
    concl.append(f"  • {r[2]}：{r[10]}客户买过/复购{r[11]}，中国仓{r[5]}")
concl += ["", "四、不用往车上搬：",
          "  • 齿轮/轴/燃油泵大件一锤子买卖，客户订了再发，不随身带",
          "  • 滤芯被国产5元件/假原厂占，除 6D8-WS24A、铃木15412-93J10 带一两个外其余按单",
          "  • 标“别带”的只有王明营一次性拿过、不具普遍性"]
concl += ["", "五、油封/O圈：驱动轴93101-28M16、车叶轴93101-30M17 为75-300走量主力，93106-09014 覆盖60-300最通用；铃木整套25700套件，详见油封专表。"]
for i, t in enumerate(concl, 1):
    c = ws2.cell(i, 1, t); c.alignment = Alignment(wrap_text=True, vertical="top")
    if t.startswith(("【", "一、", "二、", "三、", "四、", "五、")): c.font = Font(bold=True, size=11)
ws2.column_dimensions["A"].width = 110

ws3 = wb.create_sheet("口径说明")
notes = ["口径与数据来源", "",
 "1. 销售=全员6业务员153单(2026.6.11-9.4)；复购=同客户不同订单，同订单多数量只算一次",
 "2. 剔王=剔除王明营(partner156)S00115七千备货大单，避免把一次性铺货当普遍需求",
 "3. 中国仓在手=浙江远致泽昌仓(WH2/库位根1563)stock.quant实时合计，非全球；全球在手仅参考。无档案=Odoo无件号，0·中国仓缺=有档但中国仓无货(可能新加坡有)",
 "4. 适配马力：优先用你需求清单的四冲列，油封/铃木套件来自OEM爆炸图全族聚合，通用件标跨马力",
 "5. 只看四冲；齿轮/轴/燃油泵为大件一锤子买卖，滤芯受国产5元件+假原厂冲击不强推",
 "6. 销售分级：⭐好卖必带(剔王≥5客或复购≥2，卖过即证明好卖) / ⭐广覆盖通带(一件覆盖≥5个马力或跨马力通用) / 常备(2-4客) / 配套带样 / 大件按单 / 别带；库存只用来判断这次带不带得出去，不做采购补货建议",
 "7. 油封部位已按爆炸图校准：93101-28M16=驱动轴、93101-30M17=车叶(螺旋桨)轴"]
for i, t in enumerate(notes, 1):
    ws3.cell(i, 1, t).alignment = Alignment(wrap_text=True)
    if i == 1: ws3.cell(i, 1).font = Font(bold=True, size=12)
ws3.column_dimensions["A"].width = 105

out = os.path.join(OUTDIR, "2026.9.5外勤带货清单-好卖加广覆盖销售版.xlsx")
wb.save(out)
print("SAVED", out)
print("总条目", len(rows), " 好卖必带", sum(1 for r in rows if r[0]=="⭐好卖必带"),
      " 广覆盖", sum(1 for r in rows if r[0]=="⭐广覆盖通带"))
print("\n--- ⭐好卖必带 ---")
for r in [x for x in rows if x[0] == "⭐好卖必带"]:
    print(f"  {r[2].ljust(16)} {r[1].ljust(10)} {r[10]}客 复购{r[11]} 中国仓{str(r[5]):>6}")
print("\n--- ⭐广覆盖通带(前12) ---")
for r in [x for x in rows if x[0] == "⭐广覆盖通带"][:12]:
    print(f"  {r[2].ljust(16)} {r[1].ljust(10)} {str(r[4])[:40]} 中国仓{r[5]}")
