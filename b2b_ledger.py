# -*- coding: utf-8 -*-
"""B端作战台账:类型 + 试用复购状态 + 区域景气 + 信任阶段(豆包读拜访记录人工标注) + 下次动作
MANUAL 为豆包精读 res.partner.comment / crm.lead.description 后的人工判断,每次回访后更新。"""
import sys, io, os, re
from collections import defaultdict, Counter
from datetime import datetime
sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding='utf-8')
D = r"C:\Users\guobi\Doubao\chats\2026-08-16\new-chat\attivo-odoo-tools"
sys.path.insert(0, D); os.chdir(D)
from odoo.client import OdooClient
from config import ODOO_URL, ODOO_DB, ODOO_UID, ODOO_PWD
from customer_score import classify_category

od = OdooClient(ODOO_URL, ODOO_DB, ODOO_UID, ODOO_PWD)
TODAY = datetime.now().date()
PROV = {
 "福建":["福建","福州","厦门","漳州","泉州","莆田","宁德","平潭","漳浦","诏安","东山","南靖","云霄","霞浦","福鼎","溪南","梅岭","旧镇"],
 "浙江":["浙江","杭州","宁波","温州","嘉兴","湖州","绍兴","金华","衢州","舟山","台州","丽水","慈溪","苍南","龙港","瑞安","乐清","余姚","温岭","玉环"],
 "广东":["广东","广州","深圳","珠海","汕头","佛山","东莞","中山","惠州","湛江","潮汕","揭阳","汕尾","江门"],
 "江苏":["江苏","南京","苏州","无锡","常州","南通","徐州","盐城","扬州","镇江","泰州","连云港","宿迁"],
 "山东":["山东","济南","青岛","烟台","威海","日照","潍坊","东营"],
 "广西":["广西","南宁","北海","防城港","钦州"],"海南":["海南","海口","三亚","儋州"],
 "辽宁":["辽宁","大连","沈阳","丹东"],"上海":["上海"],"安徽":["安徽","合肥"],"湖北":["湖北","武汉","宜昌"],"江西":["江西","南昌","九江"],
}
def prov_of(*t):
    b=" ".join(str(x or "") for x in t)
    for p,kws in PROV.items():
        for k in kws:
            if k in b: return p
    return "其他/待补"

def ctype(nm, qty, n):
    t=str(nm)
    if re.search(r"贸易|进出口|批发|转卖|二手|回收", t): return "二道贩子"
    if re.search(r"商行|经营部|设备|机电|船艇|舷外机|动力|机械|销售|经销", t): return "经销商"
    if re.search(r"维修|修配|修理|修船|快修|服务|船外机", t): return "维修师傅"
    if qty>=8 or n>=3: return "维修师傅(行为推断)"
    return "待确认(问一句)"

# pid -> (区域景气, 信任阶段, 最佳触达时点/要点)  豆包2026-09-05精读拜访记录标注
MANUAL = {
 608:("旺","已认可待放量","常态补货:小马力齿轮月耗20-30套,要案例/要货都优先"),
 179:("淡","谨慎观望","宁德养殖崩盘没船修,齿轮没装机;10月海参季前一周再联系,现在别催"),
 301:("平","比价低意向","跟刘航月结多年,有需求按需报价即可,不投入"),
 174:("平","谨慎自测(大代理)","留了67F齿轮装自机测1-2个月,到期主动问测试结果,重点齿轮/高压油泵/叶轮"),
 744:("平","试样验证中","被国产坑过认可品质,跟进促成首单"),
 216:("平","有意深度合作","6折经销档(63D轴6折343);想建闽南中转仓可备2-3万货,谈滞销退换+材质检测背书"),
 602:("平","试样验证中","起初怀疑后已买620元,问装机结果,鲍鱼大户场景推台产"),
 298:("旺","试样验证中","诏安小马力生意最好但店里忙没装机,尽快催装机拿反馈"),
 180:("淡","有原厂渠道不迫切","有尤先生4折原厂渠道,只从稀缺件(铃木四冲)切入,维持"),
 708:("旺","已认可待放量","技术权威、习惯5-10套批量,9月新品到货深度合作"),
 810:("淡","高度认可台产","二手销量暴跌但只要台产;请他转介绍四川杨文(川黔桂流动修),油管报价单跟进"),
 166:("平","试样验证中","拿了手油泵10个/油管样,问使用反馈;内机为主挂机随缘"),
 255:("平","试样验证中","等8/15开海试60以上件,开海后跟进,自己会贴牌要留意"),
 121:("平","谨慎(只补缺口)","年采尤先生500万,只做他缺的:30高压油泵/后操拉线/新款大马力齿轮"),
 624:("平","谨慎(想换渠道)","想替换刘航、要正规增值税发票,对公报价+开票,转大客户"),
 122:("淡","试样验证中","拿了60齿轮试样,问装机;金穗65折渠道,性价比要够"),
 754:("平","观望(被假台产坑)","蓝包装台产爆齿过,发齐手册报价重建信任"),
 735:("淡(海参季旺)","低价走量试样","拿6-65折走量;现在非海参季没活不催,10月旺季前推;要极致性价比和外观辨识度"),
 156:("淡(10月旺季)","新店备货","刚开店没生意,10月旺季前跟开店首批备货,预算1万+"),
 220:("平","试样验证中","样品轴螺纹小缺陷但认可媲美原厂,问68V驱动轴装机,流水大可培育"),
 228:("平","按需采购","出租业务下滑、二手机外销广东,有维修需求时跟,不压货"),
 703:("旺","已认可复购","平潭旅游钓鱼艇,常态维护+顺带易损件"),
 719:("旺","已认可加码","精品二机、核心件只认原厂,从保养件小批量切,长期稳定"),
 184:("平","低意向(只认原厂)","说台产价高无优势,低成本挂着"),
 688:("平","谨慎(二冲为主)","85曲轴报价跟进,要日本轴承级品质"),
 697:("平","试样验证中","拿了手压泵+60齿轮试样,9月开海后问反馈"),
 168:("淡","谨慎不备货","维修活少、用到才拿,等68V(90/115)齿轮到货再推"),
 594:("平","试样验证中","拿了油管试,问反馈;水星齿轮/油泵是痛点"),
 218:("旺","试样转放量","平潭月修30波箱占70%、旺季,已拿10水泵套+2齿轮,重点跟装机趁热放量"),
 597:("淡","低意向(摆烂)","当地萧条摆烂、滤芯没用,低成本维护"),
 578:("平","试样验证中","油管试用、要300马力低压油泵,问油管反馈+报低压油泵"),
 483:("旺","试样验证中","东山外海高负荷、当场装机测试,尽快要装机结果,优质潜力"),
}

orders = od.search_read("sale.order",[["state","in",["sale","done"]],["user_id","=",18]],
    ["partner_id","date_order","amount_total","order_line"], limit=500)
lids=[l for o in orders for l in (o.get("order_line") or [])]
lines={}
for i in range(0,len(lids),200):
    for r in od.read("sale.order.line",lids[i:i+200],["product_id","product_uom_qty","name"]):
        lines[r["id"]]=r
pids=sorted({o["partner_id"][0] for o in orders})
parts={}
for i in range(0,len(pids),200):
    for r in od.read("res.partner",pids[i:i+200],["id","name","city"]):
        parts[r["id"]]=r
g=defaultdict(list)
for o in orders:
    if o.get("partner_id"): g[o["partner_id"][0]].append(o)

def first_main_cat(os_):
    o=sorted(os_,key=lambda x:str(x["date_order"]))[0]; c=Counter()
    for lid in o.get("order_line") or []:
        r=lines.get(lid)
        if r and (r.get("product_uom_qty") or 0)>0:
            pn=r["product_id"][1] if r.get("product_id") else r.get("name","")
            c[classify_category(pn)]+=1
    m={"fast":"易损件","mid":"中速件","slow":"齿轮大件"}
    return m[c.most_common(1)[0][0]] if c else ""

rows=[]
for pid, os_ in g.items():
    os_.sort(key=lambda x:str(x["date_order"]))
    ds=[str(o["date_order"])[:10] for o in os_]; amts=[o.get("amount_total",0) for o in os_]
    first,last=ds[0],ds[-1]; gap=(TODAY-datetime.strptime(last,"%Y-%m-%d").date()).days
    qty=0
    for o in os_:
        for lid in o.get("order_line") or []:
            r=lines.get(lid)
            if r and (r.get("product_uom_qty") or 0)>0: qty+=r["product_uom_qty"]
    amt=sum(amts); n=len(os_); p=parts.get(pid,{}); nm=p.get("name") or os_[0]["partner_id"][1]
    prov=prov_of(p.get("city"),nm); tp=ctype(nm,qty,n); fmc=first_main_cat(os_)
    boom,trust,timing = MANUAL.get(pid,("待补","待补","待读备注标注"))
    gaps=[(datetime.strptime(b,"%Y-%m-%d")-datetime.strptime(a,"%Y-%m-%d")).days for a,b in zip(ds,ds[1:])]
    if n==1: rep="试用未复购"
    elif any(x>=14 for x in gaps): rep="有效复购"+("·二单加码" if amts[-1]>amts[0]*1.1 else "")
    else: rep="拆单补漏·未放量"
    # 下次动作:景气×信任 优先于简单天数
    if boom.startswith("淡") and trust in("谨慎观望","试样验证中","低价走量试样","新店备货","有原厂渠道不迫切","谨慎不备货","低意向(摆烂)") and fmc=="齿轮大件":
        act="淡季下游没活/没装机,现在别催货,维持弱联系,按右侧时点再启动"
    elif trust=="已认可待放量" or trust=="已认可复购" or trust=="已认可加码" or trust=="试样转放量":
        act="趁热:要装机案例/好评,推批量备货与易损件包"
    elif trust=="试样验证中":
        act="最高优先:催装机、问运行小时/有无异常、要装机实拍(攒同地区案例)"
    elif trust.startswith("谨慎"):
        act="养:给质保包换承诺+同行案例,推动自机/熟客测试,不急着压货"
    elif trust=="比价低意向" or trust.startswith("低意向"):
        act="低成本挂着,有明确需求再报,不占跟进名额"
    elif trust=="有意深度合作":
        act="重点谈:备货仓/退换政策/材质检测背书,争取区域中转"
    else:
        act="读完备注补标注"
    rows.append([nm,prov,p.get("city",""),tp,boom,trust,n,int(qty),round(amt),first,last,gap,fmc,rep,"待问",act,timing])

# 排序:信任阶段优先级(旺地试样/认可最先)→省→gap
tprio={"试样验证中":0,"试样转放量":0,"有意深度合作":1,"已认可待放量":1,"已认可加码":1,"已认可复购":1}
prov_o={"福建":0,"浙江":1}
rows.sort(key=lambda x:(tprio.get(x[5],3), prov_o.get(x[1],2), -x[11]))

import openpyxl
from openpyxl.styles import Font,PatternFill,Alignment
wb=openpyxl.Workbook(); ws=wb.active; ws.title="B端作战台账"
head=["客户","省","城市","类型(可改)","区域景气","信任阶段","单数","累计件","累计额","首单","末单","距今天","首单品类","复购状态","品质反馈","下次动作","最佳触达时点/要点"]
ws.append(head)
hf=PatternFill("solid",fgColor="1F4E79"); ask=PatternFill("solid",fgColor="FFF2CC")
warn=PatternFill("solid",fgColor="FCE4D6"); good=PatternFill("solid",fgColor="E2EFDA")
boomfill=PatternFill("solid",fgColor="DDEBF7"); slackfill=PatternFill("solid",fgColor="F2F2F2")
for c in ws[1]: c.font=Font(bold=True,color="FFFFFF"); c.fill=hf; c.alignment=Alignment(horizontal="center")
for r in rows:
    ws.append(r); rr=ws.max_row
    ws.cell(rr,15).fill=ask
    if r[13]=="试用未复购" and r[11]>45: ws.cell(rr,14).fill=warn
    if "有效复购" in r[13] or "加码" in r[13]: ws.cell(rr,14).fill=good
    if r[4].startswith("旺"): ws.cell(rr,5).fill=boomfill
    if r[4].startswith("淡"): ws.cell(rr,5).fill=slackfill
w=[32,6,12,15,11,15,5,6,8,10,10,6,9,14,9,34,40]
for i,x in enumerate(w,1): ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=x
ws.freeze_panes="A2"

ws2=wb.create_sheet("口径与打法")
for r in [
 ["阶段定性","试用→信任转化期。客户被国产件/假台产坑过,对新台产品牌的统一决策链:小批量试样→装自机/熟客看几个月不返工→才敢批量/推客户。关键证据是'同地区×同机型×运行X小时无返工'案例,不是参数",""],
 ["",""],
 ["区域景气(读备注判断)","旺=当下有活(东山/漳浦/平潭/诏安:钓鱼旅游载客外海);淡=养殖崩盘/淡季(宁德三都澳霞浦溪南:大黄鱼鲍鱼海参行情差、船坏了不修),海参10-次年4月旺、旅游船4-10月、开海8月中/9月;平=其余",""],
 ["关键修正","兰先生齿轮没装机=宁德没船修(下游没需求)+淡季,不是品质问题也不是凉;淡季客户催复购只会招烦,要在旺季前一周启动",""],
 ["",""],
 ["信任阶段","含义","打法"],
 ["试样验证中","拿了货在等装机/看效果","最高优先,催装机问运行小时要实拍,攒案例;旺地尤其抓紧"],
 ["已认可待放量/加码","验证通过、复购或二单加码","趁热推批量备货+易损件包,发展成样板/区域背书"],
 ["谨慎自测(大代理)","要自机测1-2个月才敢推客户(季总/赖福洲/涂先生/何斌)","给质保包换+同行案例,养着,到期主动问测试结果"],
 ["比价/低意向","跟老渠道月结、只比价或摆烂","不占跟进名额,有需求再报"],
 ["",""],
 ["选品聚焦","主推:四冲60+齿轮组(90/115/150,63P/68V/69W)、高压油泵(国产1小时高温痛点)、水泵套件/叶轮、大马力轴、耐候油管、手压泵",""],
 ["","不主推:火花塞(只认NGK/原厂)、机油/汽油滤芯(与国产或原厂差价太小无利润)、60以下(国产碾压)、250/300大飞件(存量不足百台)",""],
 ["价格身位(2026-07-29大幅下调)","现行用v2026全档表(留存价格表_现行全档_20260729/),按客户折扣档报价(6/65/7/75/8折/零售/1.15),不是统一75折;调价前谈的差价作废,台产仍要比原厂便宜≥一半、和国产拉开品质差,差价太小的轴/滤芯不主推",""],
 ["增量方向","沿海卷;内陆(四川攀枝花米易150/175打波箱、广西红水河、贵州)客单价高竞争小,洪先生可转介绍流动师傅杨文;山东/江苏9月开海季节性机会",""],
 ["品质反馈取值","待问/认可/有疑虑/有质量问题(+具体);每次联系后更新",""],
]: ws2.append(r)
for c in ws2[1]: c.font=Font(bold=True)
ws2.column_dimensions["A"].width=20; ws2.column_dimensions["B"].width=70; ws2.column_dimensions["C"].width=44

out=os.path.join(os.getcwd(), "客户作战台账_B端_"+TODAY.isoformat()+".xlsx")
wb.save(out)
print("OK",len(rows),"户 ->",out)
print("景气:",Counter(r[4] for r in rows))
print("信任:",Counter(r[5] for r in rows))
