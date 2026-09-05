# -*- coding: utf-8 -*-
"""外勤装车综合建议：以桌面最新需求清单为底，叠加中国仓现货/全员销售热度/覆盖广度，
经理逻辑=摆现货、看着就挑走。复用 recommend_goods.py 的全部口径(runpy)。"""
import sys, io, os, runpy
sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding="utf-8")
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

TOOL = r"C:\Users\guobi\Doubao\chats\2026-08-16\new-chat\attivo-odoo-tools"
OUTDIR = r"C:\Users\guobi\Doubao\chats\2026-09-02\new-chat-1"
PLAN_X = r"C:\Users\guobi\Desktop\2026.9.4陈国标外勤需求清单.xlsx"

# 复用推荐脚本的全部中间结果(会顺带重存一次销售版excel,幂等无害)
g = runpy.run_path(os.path.join(TOOL, "recommend_goods.py"))
norm, ROWS = g["norm"], g["rows"]
cov_count, bigcat, is_filter = g["cov_count"], g["bigcat"], g["is_filter"]
ROWBY = {r[2]: r for r in ROWS}  # code -> [pr,cat,code,brand,hp,cnstock,gl,qall,pcall,q2,pc2,rep,plan,adv,cnname,score]

def avail_of(r):
    s = r[5]
    return int(s) if isinstance(s, (int, float)) and s > 0 else 0

# ---------- 读最新需求清单全部sheet ----------
pwb = openpyxl.load_workbook(PLAN_X, data_only=True)
plan_rows = []   # (sheet, code, cat, brand, hp4, plan)
for pws in pwb.worksheets:
    for row in pws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]: continue
        code = norm(str(row[0]))
        if not code: continue
        try: plan = int(float(row[6])) if row[6] not in (None, "") else 0
        except Exception: plan = 0
        plan_rows.append((pws.title, code, str(row[1] or ""), str(row[2] or ""),
                          str(row[4] or ""), plan, str(row[0])))
plan_codes = {x[1] for x in plan_rows}
print("清单行数", len(plan_rows), "去重件号", len(plan_codes))

# ---------- 逐行综合建议 ----------
def advise(cat, hp, plan, r):
    """返回(建议带量, 综合建议)"""
    pc2 = r[10] if r else 0; rep = r[11] if r else 0; pr = r[0] if r else ""
    avail = avail_of(r) if r else 0
    cov = cov_count(hp); wide = cov >= 5
    isbig = bigcat(cat, (r[14] if r else "")); isf = is_filter(cat, (r[14] if r else ""))
    hot = pc2 >= 5 or rep >= 2
    if avail == 0:
        stock_t = "无档案" if (r and r[5] == "无档案") else "中国仓0货"
        return 0, f"计划{plan}但{stock_t}，这次带不出，先跟仓库调/下次带"
    if isbig:
        rec = min(plan, avail) if plan else min(1, avail)
        return rec, f"大件按单、客户订了才带，计划{plan}/仓{avail}，成套齿轮记得带齐小齿+前齿+后齿"
    if isf:
        if pc2 >= 4:
            rec = min(max(plan, 5), avail); return rec, f"{pc2}客户拿过，国产竞争强，少量摆{rec}个即可"
        rec = min(plan, 3, avail) if plan else 0
        return rec, f"滤芯被国产5元件/假原厂占，计划{plan}偏多，带{rec}个摆样、要了再发"
    # 易损小件
    if hot:
        rec = max(plan, min(avail, max(6, pc2)))
        rec = min(rec, avail)
        tip = f"{pc2}客户买过" + (f"/{rep}人复购" if rep else "")
        if plan < rec: return rec, f"好卖({tip})，计划{plan}偏少、仓有{avail}，加到{rec}摆出来易被挑走"
        return min(plan, avail), f"好卖({tip})，仓{avail}够，按计划带足"
    if wide:
        rec = max(plan, min(avail, 3))
        w = "跨马力通用" if cov == 99 else f"覆盖{cov}个马力"
        return rec, f"{w}易损件，摆台师傅翻到就挑，保底{rec}个"
    rec = min(plan, avail) if plan else min(2, avail)
    return rec, ("按计划带" if plan else "没卖过、带1-2个样探需求")

out_rows = []
for sheet, code, cat, brand, hp, plan, rawcode in plan_rows:
    r = ROWBY.get(code)
    cn = r[5] if r else "无档案"; gl = r[6] if r else None
    pc2 = r[10] if r else 0; rep = r[11] if r else 0; pr = r[0] if r else "清单外未售"
    name = (r[14] if r and r[14] else cat)
    cov = cov_count(hp)
    rec, adv = advise(cat, hp, plan, r)
    out_rows.append([sheet, cat, rawcode, brand or (r[3] if r else ""), hp[:60], plan,
                     cn, gl, pc2, rep, ("跨马力" if cov == 99 else cov), rec, pr, adv, name[:30]])

# 清单外补带：好卖/广覆盖/常备 且中国仓有货 且清单没列
prow = {"⭐好卖必带": 0, "⭐广覆盖通带": 1, "常备": 2, "常备(少量)": 3, "配套带样": 4}
extra = []
for r in ROWS:
    if r[2] in plan_codes: continue
    if r[0] not in ("⭐好卖必带", "⭐广覆盖通带", "常备"): continue
    if avail_of(r) <= 0: continue
    extra.append(r)
extra.sort(key=lambda r: (prow.get(r[0], 9), -r[15]))
extra = extra[:28]

# ---------- 写 Excel ----------
wb = openpyxl.Workbook()
thin = Side(style="thin", color="BFBFBF"); bd = Border(thin, thin, thin, thin)
hfill = PatternFill("solid", fgColor="305496"); hfont = Font(bold=True, color="FFFFFF", size=10)
def style_header(ws, ncol):
    for j in range(1, ncol + 1):
        c = ws.cell(1, j); c.fill = hfill; c.font = hfont; c.border = bd
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

ws = wb.active; ws.title = "装车总表(清单逐行)"
hdr = ["清单分类", "品类", "件号(原)", "品牌", "4冲适配", "你计划带", "中国仓现货", "全球(参考)",
       "剔王客户数", "复购人数", "覆盖马力", "建议带量", "销售分级", "综合建议", "品名"]
ws.append(hdr)
recfill = {"⭐好卖必带": "C6EFCE", "⭐广覆盖通带": "FFF2CC", "常备": "DDEBF7"}
warnfill = PatternFill("solid", fgColor="FCE4D6")
for x in out_rows:
    ws.append(x)
    i = ws.max_row
    for j in range(1, 16):
        ws.cell(i, j).border = bd; ws.cell(i, j).alignment = Alignment(wrap_text=True, vertical="top")
    if x[12] in recfill:
        for j in range(1, 16): ws.cell(i, j).fill = PatternFill("solid", fgColor=recfill[x[12]])
    if x[11] == 0:  # 带不出标红
        for j in range(1, 16): ws.cell(i, j).fill = warnfill
widths = [10, 12, 20, 8, 26, 8, 9, 8, 8, 7, 8, 8, 12, 40, 16]
for j, w in enumerate(widths, 1): ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w
style_header(ws, 15); ws.auto_filter.ref = f"A1:O{ws.max_row}"

ws2 = wb.create_sheet("清单外建议补带")
h2 = ["销售分级", "品类", "件号", "品牌", "4冲适配/通用性", "中国仓现货", "剔王客户", "复购", "为什么补带"]
ws2.append(h2)
for r in extra:
    ws2.append([r[0], r[1], r[2], r[3], str(r[4])[:50], avail_of(r), r[10], r[11], r[13]])
    i = ws2.max_row
    for j in range(1, 10):
        ws2.cell(i, j).border = bd; ws2.cell(i, j).alignment = Alignment(wrap_text=True, vertical="top")
        if r[0] in recfill: ws2.cell(i, j).fill = PatternFill("solid", fgColor=recfill[r[0]])
for j, w in enumerate([13, 12, 20, 9, 30, 10, 8, 7, 42], 1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w
style_header(ws2, 9); ws2.auto_filter.ref = f"A1:I{ws2.max_row}"

# 一页纸装车指南
ws3 = wb.create_sheet("一页纸装车指南")
hot_n = [x for x in out_rows if x[12] == "⭐好卖必带"]
zero_n = [x for x in out_rows if x[11] == 0]
add_n = [x for x in out_rows if isinstance(x[5], int) and isinstance(x[11], int) and x[11] > x[5] and x[5] > 0]
guide = ["【外勤装车一页纸 — 经理逻辑：现货摆出来，师傅看着就挑走】", "",
 "一、必带：清单里客户已订的（轴/齿轮/油泵/水泵），按“建议带量”列拿；齿轮成套=小齿+前齿+后齿，缺一个整套卖不掉。",
 "二、好卖要加量（绿色，卖过证明好销，别按原计划少带）："]
for x in sorted(hot_n, key=lambda z: -z[8])[:10]:
    guide.append(f"  • {x[2]} {x[14]}：{x[8]}客户/x{x[9]}复购，计划{x[5]}→建议{x[11]}（中国仓{x[6]}）")
guide += ["", "三、摆台挑货（黄色广覆盖）：油封/断电器/水泵套件/阳极这些一件通吃多马力，桌上摊开师傅会自己翻，各带2-3个。",
          "四、清单上没有、但好卖/覆盖广且仓里有货的，见“清单外建议补带”页，顺手装上。", "",
 "五、这次带不出（红色，计划了但中国仓没货），先跟仓库说、别空答应客户："]
for x in zero_n[:12]:
    guide.append(f"  • {x[2]} {x[14]}：计划{x[5]}，{x[6]}")
guide += ["", "六、别占地方：滤芯被国产5元件/假原厂占，除6D8-WS24A、铃木15412-93J10带几个外其余2-3个摆样；",
          "  齿轮/轴/燃油泵大件一锤子买卖，客户订了才带、不囤；标“别带”的只有王明营拿过、不具普遍性。", "",
 "七、口径：销售=全员153单剔王明营大单；现货=中国仓(浙江远致泽昌)实时库存非全球；只看四冲、60-300为主。"]
for i, t in enumerate(guide, 1):
    c = ws3.cell(i, 1, t); c.alignment = Alignment(wrap_text=True, vertical="top")
    if t.startswith(("【", "一、", "二、", "三、", "四、", "五、", "六、", "七、")): c.font = Font(bold=True, size=11)
ws3.column_dimensions["A"].width = 108

out = os.path.join(OUTDIR, "2026.9.5外勤装车综合建议.xlsx")
wb.save(out)
print("SAVED", out)
print("总表", len(out_rows), "好卖加量", len(add_n), "带不出", len(zero_n), "清单外补带", len(extra))
print("\n--- 好卖需加量 ---")
for x in sorted(hot_n, key=lambda z: -z[8]):
    print(f"  {x[2].ljust(18)} 计划{x[5]} -> 建议{x[11]} (仓{x[6]}, {x[8]}客)")
print("\n--- 带不出 ---")
for x in zero_n: print(f"  {x[2].ljust(18)} 计划{x[5]} {x[6]}")
