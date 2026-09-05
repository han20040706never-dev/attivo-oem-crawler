# -*- coding: utf-8 -*-
"""油封/O圈/密封套件 带货推荐 v2：只看中国仓有货；没卖过但适配60-300马力也推荐"""
import sys, io, os, re, json
sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding="utf-8")
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TOOL = r"C:\Users\guobi\Doubao\chats\2026-08-16\new-chat\attivo-odoo-tools"
OUTDIR = r"C:\Users\guobi\Doubao\chats\2026-09-02\new-chat-1"
SRC = os.path.join(OUTDIR, "2026.9.5油封O圈带货推荐-机型部位对照.xlsx")
OUT = os.path.join(OUTDIR, "2026.9.5油封O圈带货推荐-中国仓有货60-300版.xlsx")
WANG = 156
LO, HI = 60, 300
# 部位专有名词纠错(用户确认, 优先于源表标注)
import parts_kb as KB  # 统一业务知识库(部位纠错/碳刷组合/老款/葡语品类)
def fix_pos(part, pos):
    return KB.fix_pos(part, pos)

def norm(c):
    """件号归一单一真相源 cn_stock.norm（保留材质 AL/ZN 与 -W/H）。"""
    import cn_stock
    return cn_stock.norm(c)
def parse_hp(s):
    out = set()
    for x in re.findall(r"\d+", str(s or "")):
        v = int(x)
        if 2 <= v <= 350: out.add(v)
    return out

# 1. 销售(全员,剔王)
sd = json.load(open(os.path.join(TOOL, "sales_data.json"), encoding="utf-8"))
op = {o["id"]: o["partner"] for o in sd["orders"]}
pcode = {int(pid): v.get("code") or "" for pid, v in sd["products"].items()}
S = {}
def bk(k): return S.setdefault(k, {"qty": 0.0, "ord": set(), "cust": set(), "ordx": set(), "custx": set()})
for ln in sd["lines"]:
    code = pcode.get(ln["product"]) or (re.search(r"\[([^\]]+)\]", ln.get("name", "")).group(1) if re.search(r"\[([^\]]+)\]", ln.get("name", "")) else "")
    k = norm(code)
    if not k: continue
    oid = ln["order"]; pid = op.get(oid); b = bk(k); q = ln.get("qty") or 0
    b["qty"] += q; b["ord"].add(oid); b["cust"].add(pid)
    if pid != WANG: b["ordx"].add(oid); b["custx"].add(pid)

# 2. 中国仓库存
_cf = os.path.join(TOOL, "_cn_stock_cache.json")
ST = json.load(open(_cf, encoding="utf-8"))["ST"] if os.path.exists(_cf) else {}
if not ST:
    print("WARN _cn_stock_cache.json 缺失/为空：先跑 recommend_goods.py 建缓存，否则推荐页为空")
def stock(k):
    v = ST.get(norm(k))
    return (round(v.get("cn", 0) or 0, 1), round(v.get("gl", 0) or 0, 1)) if v else (0.0, 0.0)

# 缓存0实时复核: 只复核油封/O圈/密封类(931/932/0928开头)缓存判0的件, 根治缓存过时把有货误判0而漏推荐
try:
    import cn_stock as _CN
    _z = [k for k, v in ST.items() if (v.get("cn", 0) or 0) <= 0 and re.match(r"^(931|932|0928)", k)]
    _lv = _CN.live_qty(_z)
    _fixed = 0
    for _k, _q in _lv.items():
        if _k in ST and _q > 0:
            ST[_k]["cn"] = _q; _fixed += 1
    print(f"密封件缓存0实时复核 {len(_z)} 件, 纠正为有货 {_fixed} 件")
except Exception as _e:
    print("实时复核跳过:", str(_e)[:80])

# 3. 读原表 + 用机型对照聚合每件号适配马力
wb0 = load_workbook(SRC, read_only=True, data_only=True)
master = [list(r) for r in wb0["油封件号主数据"].iter_rows(min_row=2, values_only=True) if r[0]]
xwalk = [list(r) for r in wb0["机型x部位对照"].iter_rows(min_row=2, values_only=True) if r[3]]
wb0.close()
part_hp = defaultdict(set)
for r in xwalk:
    part_hp[norm(r[3])] |= parse_hp(r[1])

ADVICE = {
 "⭐好卖必带": "剔王后多客户/复购、中国仓有货，优先带足",
 "卖过可带": "剔王后有客户买过、有货，适量带",
 "新品可带(60-300)": "没卖过但适配60-300马力、中国仓有货，带着铺给中大马力客户",
 "铃木整套(按单)": "铃木齿轮箱密封整套，按对应机型带",
 "卖过·缺货": "卖过但中国仓0，登记需求/调货，不带现货",
 "缺货暂不带": "中国仓0现货，不占带货名额",
 "暂不带(仅小马力)": "只适配60以下小马力，非目标区间",
 "暂不带(老款)": "老款机型油封，不当新品带"}
REC_GRADES = ["⭐好卖必带", "卖过可带", "新品可带(60-300)", "铃木整套(按单)"]
ORDER = {g: i for i, g in enumerate(REC_GRADES)}
GFILL = {"⭐好卖必带": "C6EFCE", "卖过可带": "E2EFDA", "新品可带(60-300)": "DDEBF7", "铃木整套(按单)": "FCE4D6"}

def grade(part, hp, cn):
    if KB.is_old(part): return "暂不带(老款)"
    b = S.get(norm(part)); nx = len(b["ordx"]) if b else 0; cx = len(b["custx"]) if b else 0
    inrng = any(LO <= h <= HI for h in hp)
    suz = part.startswith("25700")
    if cn <= 0:
        if (nx >= 1 or cx >= 1): return "卖过·缺货"
        return "缺货暂不带"
    if nx >= 2 or cx >= 2: return "⭐好卖必带"
    if nx >= 1 or cx >= 1: return "卖过可带"
    if suz: return "铃木整套(按单)" if inrng else "暂不带(仅小马力)"
    return "新品可带(60-300)" if inrng else "暂不带(仅小马力)"

# 4. 主数据计算
mrows = []
for r in master:
    part, _pos0, brand, fit, cov, cname = r[0], r[1], r[2], r[3], int(r[4] or 0), r[5]
    pos = fix_pos(part, _pos0)
    hp = part_hp.get(norm(part)) or parse_hp(fit)
    hp_in = sorted(h for h in hp if LO <= h <= HI)
    cn, gl = stock(part); b = S.get(norm(part))
    g = grade(part, hp, cn)
    mrows.append([g, part, pos, brand, fit, cov, "/".join(map(str, hp_in)) or "—", cn, gl,
                  round(b["qty"], 1) if b else 0, len(b["ord"]) if b else 0, len(b["cust"]) if b else 0,
                  len(b["custx"]) if b else 0, cname or "", ADVICE[g]])

# 5. 样式
HDR = PatternFill("solid", fgColor="1F4E78"); HF = Font(color="FFFFFF", bold=True, size=10)
thin = Side(style="thin", color="BFBFBF"); BD = Border(thin, thin, thin, thin)
def head(ws, n):
    for c in range(1, n+1):
        x = ws.cell(1, c); x.fill = HDR; x.font = HF
        x.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); x.border = BD
    ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{get_column_letter(n)}1"; ws.row_dimensions[1].height = 30
def body(ws, rows, w):
    for ri, row in enumerate(rows, 2):
        for ci, v in enumerate(row, 1):
            x = ws.cell(ri, ci, v); x.border = BD
            x.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center" if ci != len(row) else "left")
    for i, ww in enumerate(w, 1): ws.column_dimensions[get_column_letter(i)].width = ww

wb = Workbook(); wb.remove(wb.active)
# Sheet1 带货推荐：只中国仓有货 & 属推荐等级
rec = [m for m in mrows if m[7] > 0 and m[0] in REC_GRADES]
rec.sort(key=lambda m: (ORDER[m[0]], -m[5], -m[9]))
ws1 = wb.create_sheet("⭐带货推荐(中国仓有货)")
h1 = ["推荐等级", "件号", "部位", "品牌", "适配马力(原文)", "覆盖马力数", "60-300适配", "中国仓在手", "全球在手",
      "全员销量", "订单数", "客户数", "剔王客户", "中文名/规格", "带货建议"]
ws1.append(h1); body(ws1, rec, [15,16,11,8,18,8,14,9,8,8,7,7,8,20,30])
for ri, m in enumerate(rec, 2):
    if m[0] in GFILL: ws1.cell(ri,1).fill = PatternFill("solid", fgColor=GFILL[m[0]]); ws1.cell(ri,1).font = Font(bold=True)
head(ws1, len(h1))

# Sheet2 主数据全量
ws2 = wb.create_sheet("油封件号主数据(全量)")
ms = sorted(mrows, key=lambda m: (ORDER.get(m[0], 8), m[2], -m[7]))
ws2.append(h1); body(ws2, ms, [15,16,11,8,18,8,14,9,8,8,7,7,8,20,26])
for ri, m in enumerate(ms, 2):
    if m[0] in GFILL: ws2.cell(ri,1).fill = PatternFill("solid", fgColor=GFILL[m[0]])
head(ws2, len(h1))

# Sheet3 机型x部位对照(只60-300, 更新库存销量)
ws3 = wb.create_sheet("机型x部位对照60-300")
h3 = ["品牌","马力","部位","件号","中文名/规格","中国仓在手","全球在手","销量","剔王客户","推荐等级"]
ws3.append(h3); xr = []
for r in xwalk:
    brand, hpv, _posx, part, cname = r[0], r[1], r[2], r[3], r[4]
    pos = fix_pos(part, _posx)
    hps = parse_hp(hpv)
    if not any(LO <= h <= HI for h in hps): continue
    cn, gl = stock(part); b = S.get(norm(part))
    g = grade(part, part_hp.get(norm(part)) or hps, cn)
    xr.append([brand, hpv, pos, part, cname or "", cn, gl, round(b["qty"],1) if b else 0, len(b["custx"]) if b else 0, g])
body(ws3, xr, [8,7,12,16,22,10,9,8,8,15])
for ri, x in enumerate(xr, 2):
    if x[9] in GFILL: ws3.cell(ri,10).fill = PatternFill("solid", fgColor=GFILL[x[9]])
head(ws3, len(h3))

# Sheet4 已售密封件
ws4 = wb.create_sheet("已售密封件验证")
h4 = ["件号","全员销量","订单数","客户数","剔王订单","剔王客户","中国仓在手","全球在手","中文名"]; ws4.append(h4); sr=[]
seal = re.compile(r"^(93101|93102|93106|93210|93211|932-|09283|25700|0928|93110)", re.I)
for k, b in S.items():
    if seal.search(k):
        cn, gl = stock(k); sr.append([k, round(b["qty"],1), len(b["ord"]), len(b["cust"]), len(b["ordx"]), len(b["custx"]), cn, gl, ST.get(norm(k),{}).get("name","")])
sr.sort(key=lambda x:-x[1]); body(ws4, sr, [18,9,8,8,9,9,10,9,24]); head(ws4, len(h4))

# Sheet5 口径
ws5 = wb.create_sheet("口径说明")
notes = [["油封/O圈/密封套件 带货推荐（中国仓有货 · 60-300马力版）",""],
 ["本版原则","只推荐中国仓(WH2/1563)有现货的；没卖过但适配60-300马力的油封也推荐；缺货/仅小马力不占带货名额"],
 ["销售口径","全员订单(2026-06至09-04,153单)；王明营partner156一次性大单在'剔王客户'剔除，其独买不算好卖"],
 ["库存口径","中国仓=库位根1563的internal合计；'全球在手'仅对照，不用qty_available(含新加坡)"],
 ["⭐好卖必带","剔王后≥2客户或≥2订单且有货"],
 ["卖过可带","剔王后有客户买过且有货"],
 ["新品可带(60-300)","没卖过，但适配马力落在60-300且中国仓有货，铺给中大马力客户"],
 ["复购口径","同订单多数量算一次，不同订单才算复购"],
 ["部位","28M16驱动轴/30M17车叶轴/93102曲轴/93106驱动轴/93210·93211 O型圈；25700铃木整套；W/H带壳不合并"]]
for r in notes: ws5.append(r)
ws5.column_dimensions["A"].width=18; ws5.column_dimensions["B"].width=92; ws5["A1"].font=Font(bold=True,size=12)
for ri in range(2,len(notes)+1):
    ws5.cell(ri,1).font=Font(bold=True); ws5.cell(ri,2).alignment=Alignment(wrap_text=True,vertical="center")

wb.save(OUT)
from collections import Counter
print("SAVED", OUT); print("推荐(有货)条数:", len(rec)); print(Counter(m[0] for m in mrows))
print("\n=== 中国仓有货·推荐带货清单 ===")
for m in rec:
    print(f"{m[0]:<14}{m[1]:<15}{m[2]:<9}60-300[{m[6]}] 仓{m[7]:>5} 销{m[9]:>4} 剔王客{m[12]}")
